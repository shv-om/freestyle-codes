"""
ye program excel file k 2nd column m name se regex bnake search destination
me saari files k names m dhoondta h regex ko jisme milta h us row k aakhri
column m found likh deta h or agar plant k name me "subsp. var. ya syn." to
files me bhi "subsp. ..." se hi dhoondta h. ye confirm krne k lliye ki subspecies
vali files k aakhri column me hi found with subspecies lage.

ye program excel file modify krta h to 'Shivam' se permission liye bina is
program ko run na kre.
Thank you

"""
import os, re
import openpyxl

foundfiles = open('file2excel list.txt', 'w')
foundfiles.close()
filelist = []

search_fol = input('Please input the path of folder where to search: ')

def file2excel():

	workbook = openpyxl.load_workbook('test.xlsx')
	sheetlist = workbook.sheetnames		#taking sheetname as list

	namerex = re.compile(r'''(
							([A-Za-z]+)
							(\s)+
							([A-Za-z]+)
							(.*)
							)''', re.VERBOSE | re.I)

	subnamerex = re.compile(r'''(
							([A-Za-z]+)
							(\s)+
							([A-Za-z]+)
							(\s)+
							(.*(subsp|subs|sub|syn|var)\.)*
							(\s)*
							([A-Za-z]*)
							(.*)
							)''', re.VERBOSE | re.I)

	for n in sheetlist:
		sheet = workbook[n]
		print(sheet)
		cols = sheet.max_column + 1
		cell_found_status = sheet.cell(row = 1, column = cols)
		cell_found_status.value = 'Image Status'

		for i in range(2, sheet.max_row + 1):

			#agar subsp... milta h plant name me to subnamerex se group create honge
			#or phir files me bhi subsp... hona jruri h taaki confirm ho jaye ki file usi plant ki h
			if 'subsp.' in sheet.cell(row = i, column = 2).value.lower() or 'syn.' in sheet.cell(row = i, column = 2).value.lower() or 'var.' in sheet.cell(row = i, column = 2).value.lower():
				for groups in subnamerex.findall(sheet.cell(row = i, column = 2).value.lower()):
					for rootfolder, subfolder, filenames in os.walk(search_fol):
						for filename in filenames:
							print(rootfolder+', '+ filename)
							print(sheet.cell(row = i, column = 2).value.lower())
							#print(filename)
							#print(groups[6]+', '+groups[8]+', '+groups[9])
							if groups[1].lower() in filename.lower() and groups[3].lower() in filename.lower() and groups[6].lower() in filename.lower() and groups[8].lower() in filename.lower():
								print('subs')
								filelist.append(sheet.cell(row = i, column = 2).value)
								foundfiles = open('file2excel list.txt', 'a')
								foundfiles.write(sheet.cell(row = i, column = 2).value +': Found with subspecies\n')
								#writing excel file
								#write_cell = sheet.cell(row = i, column = cols)
								#write_cell.value = 'Found with subspecies'
								break
							else:
								print('not subs')
								#write_cell = sheet.cell(row = i, column = cols)
								#write_cell.value = 'Not Found'


			#agar subsp... nhi miltah to ye run hoga taki bs genus or species k naam se search ho
			#pr ye files k name m bhi subsp.. na ho ye bhi confirm krta h
			else:
				for groups in namerex.findall(sheet.cell(row = i, column = 2).value.lower()):
					#namelist.append(groups[1] +' '+ groups[3])
					for rootfolder, subfolder, filenames in os.walk(search_fol):
						for filename in filenames:
							print(rootfolder+', '+ filename)
							print(sheet.cell(row = i, column = 2).value.lower())
							#print(filename)
							if 'subsp.' not in filename.lower() and 'var.' not in filename.lower() and 'syn.' not in filename.lower():
								if groups[1].lower() in filename.lower() and groups[3].lower() in filename.lower():
									print('only genus')
									filelist.append(sheet.cell(row = i, column = 2).value)
									foundfiles = open('file2excel list.txt', 'a')
									foundfiles.write(sheet.cell(row = i, column = 2).value +': Found with only genus and species\n')
									#writing excel file
									#write_cell = sheet.cell(row = i, column = cols)
									#write_cell.value = 'Found'
									break
								else:
									print('not genus')
									#write_cell = sheet.cell(row = i, column = cols)
									#write_cell.value = 'Not Found'

			print(filelist)
			if sheet.cell(row = i, column = 2).value in filelist:
				print('yes in filelist')
				write_cell = sheet.cell(row = i, column = cols)
				write_cell.value = 'Found'

	workbook.save('test.xlsx')		#saving the excel file

file2excel()
